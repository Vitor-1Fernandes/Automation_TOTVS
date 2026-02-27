from openpyxl import load_workbook
import pyautogui
import time
import keyboard
import pygetwindow as gw

planilha = load_workbook('Produtos.xlsx', data_only=True)  

pagina = planilha['Sheet1']

produto = []
quantidade = []
armazen = []
documento = "18A9782A"

linha = 0
for dados in pagina.iter_rows(values_only=True):
    if linha == 0:
        pass
    else:
        produto.append(dados[0])
        quantidade.append(dados[1])
        armazen.append(dados[2])
    linha = linha + 1

def clicar(vezes, tecla):
    for vez in range(vezes):
        time.sleep( 0.1) 
        keyboard.send(tecla)
     
time.sleep(1) 
keyboard.send('windows')

time.sleep(1)
pyautogui.write("Totvs")

time.sleep(1)
keyboard.send('enter')

time.sleep(1)
keyboard.send('tab')
pyautogui.write("admin")

time.sleep(1)
keyboard.send('tab')
pyautogui.write("sccp@1910")

keyboard.send('tab')
keyboard.send('enter')

for i in range(len(produto)):
    if i == 0:
        time.sleep(3)
        clicar(3, 'tab')
        pyautogui.write(str(produto[i]))    
    else: 
        time.sleep(3)
        clicar(1, 'tab')
        pyautogui.write(str(produto[i]))
    
    clicar(1, 'tab')
    pyautogui.write(str(quantidade[i]))

    clicar(5, 'tab')
    pyautogui.write(str(armazen[i]))

    clicar(12, 'tab')
    pyautogui.write(str(documento))

    clicar(7, 'tab')
    keyboard.send('enter')


